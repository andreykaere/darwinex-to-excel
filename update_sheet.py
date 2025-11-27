#!/usr/bin/python3.5

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

from update_info import *

file = 'DWX-MACRO.xlsx'
WB = load_workbook(filename=file)
SHEET = WB.active

FIELDS1 = FIELDS2 = FIELDS3 = ["ex", "anos", "la", "equity", "rentabilidad",
                               "d_score", "divergencia"]

REDFILL = PatternFill(start_color='F4B183',
                   end_color='F4B183',
                   fill_type='solid')
GREENFILL = PatternFill(start_color='A9D18E',
                   end_color='A9D183',
                   fill_type='solid')

#WHITEFILL = PatternFill(start_color='FFFFFF',
#                   end_color='FFFFFF',
#                   fill_type='solid')

def update_column(col):
    column = dict()
    row = 3
    cord = col + "3"

    while SHEET[cord].value != None:
        column[SHEET[cord].value] = (col, row)  
        row += 1
        cord = col + str(row)

    return column



def get_color(field, value):
    benchmark = None

    if field == "anos":
        benchmark = 4
    if field == "d_score":
        benchmark = 65
    if field == "divergencia":
        benchmark = 0
    if field == "la":
        benchmark = 7
    if field == "equity":
        benchmark = 8000
    if benchmark == None:
        return "WHITE"

    if value >= benchmark:
        return "GREEN"
    else:
        return "RED"



def next(char):
    order = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ']

    return order[order.index(char) + 1]


def fill_cell(name, cords, FIELDS):
    (char, num) = cords
    char = next(char)
    data = INFO[name]
        
    for field in FIELDS:
        cord = char + str(num)
        value = data[field]
        color = get_color(field, value)
        print(cord, field, value, color)
        SHEET[cord] = value

        if field in ['divergencia', 'rentabilidad']:
            SHEET[cord] = str(value) + '%'

        if field == 'equity':
            if value >= 8000:
                SHEET[cord] = '+'
            else:
                SHEET[cord] = '-'

        if color != "WHITE":
            SHEET[cord].fill = eval(color + "FILL")
        char = next(char)


def fill_category(category, FIELDS):
    for name in list(category.keys()):
        fill_cell(name, category[name], FIELDS)


def fill_sheet():
    fill_category(INVIRTIENDO, FIELDS1)
    fill_category(OBSERVACION, FIELDS2)
    fill_category(OBSERVACION_SIN_EXP, FIELDS3)


INVIRTIENDO         = update_column("B")
OBSERVACION         = update_column("L")
OBSERVACION_SIN_EXP = update_column("V")


ALL_NAMES = set(list(INVIRTIENDO.keys()) + \
                list(OBSERVACION.keys()) + \
                list(OBSERVACION_SIN_EXP.keys()))

for name in ALL_NAMES:
    get_data(name)

fill_sheet()
WB.save(filename=file)
